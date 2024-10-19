from openpyxl import load_workbook

def filter_groups(input_file, results_file):
    # Load the workbook with openpyxl
    wb = load_workbook(input_file)
    sheet = wb.active
    
    #most of the time the index of the job title is 2 or 3
    print("Which index does the job title column has (most of the time it is 2 or 3)?")
    print("Remember that python index's starts with 0, so A=0, B=1, C=2....")
    print("And that you will get an error if you fill in something else then a integer.")

    while True:
        job_title_row = input("the index of the column is:")

        # Try to convert the input to an integer
        try:
            job_title_row = int(job_title_row)
            print("You entered an integer, well done! \n")
            break
        except ValueError:
            print("The value is not an integer, try again \n")

    # Define the strings to check
    to_check_cfo = ["cfo", "chief financial director", "chief financial officer", "chief financial", \
        "chief finance", "financial officer", "finance officer"]
    to_check_ceo = ["ceo", "chief executive officer", "owner"]
    to_check_product = ["hop", "head of product", "vp of product", "cpo", "chief product officer", \
        "vice president of product", "Chief Product", "head of ai", "head of Digital Product", \
        "head of it", "head of software", "head of tech", "product", "head product"]
    to_check_cto = ["cto", "chief technology officer", "chief technical officer", \
        "chief data", "chief digital & technology officer", "chief technology", "technology officer"]
    to_check_HoE = ["head"]
    to_check_VPoE = ["vp", "vice president"]
    to_check_MoE = ["engineering manager"]
    to_check_coo = ["coo", "chief operating officer", "chief operations officer", "chief operations", \
                    "chief operating", "operations officer", "chief operations", "chief operation", \
                    "chief operational officer", "chief operationals", "operating officer", "operations officer"]
    to_check_cio = ["cio", "chief information officer", "Bereichsleiter Informationstechnologie", \
        "chief information"]
    to_check_ciso = ["ciso", "chief information security officer"]
    to_check_cso = ["cso", "chief sales officer"]
    # this is an extra check to delete bad leads
    to_check_trash = ["cto office", "Clinical", "chief people officer", "chief of staff", "Account manager for ZZP in construction & Engineering",
    "Advanced App Engineering Assoc Manager @ Accenture LS Netherlands",
    "Advanced App Engineering Associate Manager",
    "AOC Engineering & Technology Manager / MOT Supervisory Board Chairman",
    "Applicatie Manager for 'Engineering'",
    "Area Manager, Network Engineering",
    "Area Sales Manager, System Engineering",
    "Assistent manager / teamleider mechanical Engineering",
    "Assistent Manager Maintenance & Engineering",
    "Associate Manager Engineering",
    "Associate manager external manufacturing engineering",
    "Business domain manager for IT Component Services at Engineering & Maintenance",
    "Business Lead / Senior Manager - Data Engineering",
    "Business Manager - Engineering Interim Division",
    "Business Manager Engineering Projects",
    "Business Program Manager - Industry Software Engineering",
    "Business Unit Manager Automotive & Engineering",
    "Business Unit Manager Digital Engineering Experiences",
    "Business Unit Manager Industrie & Engineering",
    "Business Unit Manager JMEC Engineering",
    "Business Unit Manager Technology - Mechanical Engineering",
    "Category Manager Engineering & Steel constructions SMD",
    "CHIME System Engineering and Operations Manager",
    "Cloud Strategy and Engineering Senior Manager",
    "Cluster Manager CS DUV Engineering - Bottom&Layout",
    "Cluster Quality Manager Business Line Applications Development & Engineering",
    "Data Center Engineering Operations Cluster Manager",
    "Data Engineering Associate Manager",
    "Delivery Manager CDC Security Engineering",
    "Delivery Manager Engineering and Build",
    "DELMIA RelOps Engineering Senior Manager",
    "Department Manager Applications & Engineering",
    "Department Manager at Quantum and Computer Engineering, TU Delft",
    "Department Manager CSA & Piping Engineering",
    "Department Manager Design HSE and Lead EMS/BMS Harmonization Safety Engineering",
    "Department Manager Electrical & Instrumentation Engineering",
    "Department Manager Electronics & Electrical Engineering",
    "Department Manager Engineering",
    "Department Manager Engineering & Projectmanagement at Siemens",
    "Department Manager Engineering & Technical Services",
    "Department manager Hybrid Cloud engineering",
    "Department Manager maintenance & engineering",
    "Department Manager Mechanical & HSE Engineering",
    "Department Manager Piping Engineering & Design",
    "Department Manager Process & Safety Engineering",
    "Deputy Research Manager @ ACE - Advanced Computing Engineering",
    "Design and Engineering Quality Manager",
    "Design Engineering & Architecture Line Manager - Process and Cleanliness Control EUV",
    "Design Engineering and Architecture Line Manager",
    "Design Engineering Quality Manager",
    "Design Engineering Quality Program Manager",
    "Discipline Manager Reservoir Engineering Conventional Oil & Gas",
    "Duty Maintenance Manager- Engineering & Maintenance",
    "Dy General Manager - Supply Chanin Engineering",
    "Education Manager BSc Mechanical Engineering UT-VUE",
    "Educational manager at Engineering institute",
    "Educational Manager Electrical Engineering",
    "Engagement Manager at Van Oord Civil Engineering - Marine Ingenuity",
    "Engineering & Maintenance Manager",
    "Engineering & Maintenance Manager at Enza Zaden",
    "Engineering & Process Innovation Manager",
    "Engineering & Projects Manager",
    "Engineering and Delivery Manager Energy Infrastructure Solutions",
    "Engineering and Service Manager",
    "Engineering Change Control Manager",
    "Engineering Change manager",
    "Engineering Data and Application Discipline Manager",
    "Engineering Operations Governance Manager",
    "Engineering Operations Lead - Manager/Consultant",
    "Engineering Outsourcing Manager",
    "Engineering Program Manager",
    "Engineering Program Manager Foods and Beverages",
    "Engineering Resolution Manager",
    "Engineering Resource Manager",
    "Engineering Service Manager",
    "Engineering Services Manager",
    "Engineering- Release Manager",
    "Engineering/NPI Manager",
    "Environment, Health and Safety Manager & PM Engineering",
    "Expert IT Engineering, CMO BI team member",
    "Faculty IT Manager Applied Sciences and Industrial Design Engineering",
    "Fourth year Aerospace Engineering Student @ TU Delft | President",
    "GCS Team Manager ARIS Engineering and Technical Consulting Services",
    "General Manager Sales Engineering",
    "General Manager Sales Engineering Propulsion",
    "Geotechnical Engineer and Manager of Engineering",
    "Global Engineering Operations Manager",
    "Global Enterprise Lead and Engineering Hiring Manager",
    "Global Information Manager - Manufacturing and Engineering",
    "Global Manager Advanced Quality Engineering",
    "Global Manager Cost Engineering",
    "Global Manager Process Engineering",
    "Global Projects Engineering Capability Manager",
    "Global Sales Engineering and Technical Support Manager",
    "Global SPS Manager: Logistic Cost Engineering",
    "Global Sr. Manager Engineering - M&A Industrial BU",
    "Group Discipline Manager",
    "Group manager Development and engineering EUV Metrology",
    "Group Manager Engineering & Maintenance",
    "Group Manager EUV Electronics Development - Industrialization Engineering",
    "Hub Manager Engineering Europe-West",
    "IDT Manager Offshore Wind Projects and Engineering",
    "Improvement Manager Engineering",
    "Infrastructure & Engineering Services Manager",
    "Interim Sr. Manager Facilities & Engineering",
    "IT Manager - Azure Platform Engineering",
    "Jr Manager Data Engineering & Analytics",
    "Line Manager Design Engineering and Architecture",
    "Line manager System Solution Engineering",
    "Maintenance Manager Mobility Network Engineering",
    "Manager - Business Value Engineering",
    "Manager - Process Engineering / Commissioning",
    "Manager Aftermarket Engineering",
    "Manager Aftersales & Engineering a.i. | Vestigingsmanager Amsterdam & Geleen a.i.",
    "Manager CAPEX Projecten Formulation Dept Engineering",
    "Manager consultant engineers Loss Prevention and Risk Management / Group Manager Field Engineering",
    "Manager Controls Engineering",
    "Manager Controls Engineering at BDR Thermea Group B.V.",
    "Manager Controls Retrofit Engineering",
    "Manager Cost Engineering",
    "Manager Cost Engineering proposals",
    "Manager Design & Engineering International Network Services",
    "Manager Engineering, Gracenote Sports bij Gracenote",
    "Manager Enterprise Engineering Analytics Enablement Team",
    "Manager Environmental Engineering Department",
    "Manager Expert Field Service Engineering Food Processing & Packaging",
    "Manager Faculty Engineering, Design and Computing | Lead Operations & Business Control",
    "Manager Field Engineering",
    "Manager FMI Engineering",
    "Manager Growth Engineering",
    "Manager GTL Engineering",
    "Manager ICT & Logistical Engineering",
    "Manager IDE4P: technisch onderwijs voor professionals bij Institute for Design & Engineering",
    "Manager Identity Engineering",
    "Manager Identity Engineering - Digital Identity Solution Architect",
    "Manager Industrialization Engineering, Vacuum Systems & Fluid and temperature Control",
    "Manager Installation & Engineering",
    "Manager Installation engineering Real Estate & Facility Management department",
    "Manager Integrity and Maintenance Engineering",
    "Manager Launch Engineering",
    "Manager layout engineering",
    "Manager Life Sciences & Pharmaceutical Engineering",
    "Manager Maintenance & Engineering",
    "Manager Maintenance & Reliability Engineering",
    "Manager Maintenance & Reliability Engineering a.i.",
    "Manager Maintenance and Engineering",
    "Manager maintenance Development & RAMS/LCC Engineering",
    "Manager maintenance en engineering",
    "Manager Maintenance Engineering",
    "Manager maintenance engineering & planning",
    "Manager Maintenance Engineering Netherlands",
    "Manager Maintenance, Engineering & EHS",
    "Manager Maintenance, Engineering & Facilities",
    "Manager Maintenence and Engineering",
    "Manager mechanical engineering & standardisation",
    "Manager MfS OpEx Engineering Services",
    "Manager of control software engineering group",
    "Manager of Solution Engineering, Growth Solutions",
    "Manager Office Engineering & Training",
    "Manager Ontwikkeling, Engineering & Acquisitie / Management Team @Ballast Nedam Zuid",
    "Manager Operation Pipeline Engineering",
    "Manager Operational Engineering",
    "Manager Operations & Engineering",
    "Manager Operations & Engineering Sollas Holland",
    "Manager Operations and Engineering",
    "Manager Operations, projects en engineering",
    "Manager Order Engineering",
    "Manager Piping Engineering",
    "Manager Planning - Systems Engineering",
    "Manager Planning & Engineering",
    "Manager Planning & Engineering - Sort Technology Systems",
    "Manager Planning & Engineering O- perations Technology Engineering",
    "Manager Planning & Engineering Properties",
    "Manager Planning Engineering",
    "Manager Portfolio Deployment & Training, Industrial Engineering Solutions, Europe",
    "Manager Portfoliomanagement & collections Engineering",
    "Manager Private Cloud Engineering",
    "Manager Proces Engineering",
    "Manager Proces Engineering engine assembly en test",
    "Manager Process and Engineering",
    "Manager Process Engineering",
    "Manager Process Engineering (CNG, LNG and Hydrogen)",
    "Manager Process Engineering & LSS",
    "Manager Process Engineering and Electrolysis Subject Matter Expert",
    "Manager Process Engineering Department",
    "Manager Process Engineering TE&P",
    "Manager Process Quality Engineering",
    "Manager Project Engineering",
    "Manager Project Engineering - MT Lid",
    "Manager Project Engineering & Line Solutions",
    "Manager Project Engineering TA",
    "Manager Project Management & Engineering",
    "Manager Project management & System Test Engineering Advanced Systems EUV",
    "Manager project management en engineering",
    "Manager Project management & System Test Engineering Advanced Systems EUV",
    "Manager project management en engineering",
    "Manager Project Mold & Tool Engineering",
    "Manager Projectbeheersing / Systems Engineering",
    "Manager Projecten & Engineering",
    "Manager Projects & Engineering",
    "Manager projects engineering & IT",
    "Manager Projects Engineering & Maintenance",
    "Manager Proposal Engineering",
    "Manager Quality -System Engineering-",
    "Manager Quality Engineering - BTO Customer",
    "Manager Quality Engineering & Support",
    "Manager RAN Engineering",
    "Manager regional engineering",
    "Manager Reliability Engineering",
    "Manager Risk Engineering",
    "Manager Sales / Innovation & Engineering",
    "Manager Sales & Applications Engineering",
    "Manager sales en design engineering",
    "Manager sales engineering",
    "Manager Sales Engineering - Benelux",
    "Manager Sales Engineering @AppDynamics (Cisco) Benelux & Denmark",
    "Manager Sales Engineering & Support",
    "Manager sales engineering and aftersales Component sales",
    "Manager Sales Engineering International",
    "Manager Security Engineering",
    "Manager Service desk & IT engineering",
    "Manager Service Engineering",
    "Manager service-engineering",
    "Manager Simulation Engineering",
    "Manager Site Engineering & Facilities",
    "Manager Solution Engineering",
    "Manager Solutions Engineering",
    "Manager Support & Engineering",
    "Manager System Engineering Air Traffic Management at Saab Nederland B.V.",
    "Manager System Engineering: Applications - Metrology & Data",
    "Manager Systems Engineering",
    "Manager Systems Engineering - Channel / Telco",
    "Manager Systems Engineering - Mid Enterprise & Territory",
    "Manager T&H Engineering",
    "Manager Technical department and engineering",
    "Manager Technical Documentation and Engineering Support",
    "Manager Technical Sales Support & Engineering",
    "Manager technical support & engineering",
    "Manager Technical Support Engineering Heavy Lifting",
    "Manager Technisch Expertisecentrum/Engineering",
    "Manager Technische Dienst & Engineering",
    "Manager Technology & Engineering",
    "Manager Technology & Engineering - Connectivity & Security",
    "Manager tender engineering (Subsea cables)",
    "Manager Test & Verification Engineering",
    "Manager Test Development Engineering",
    "Manager Test Engineering",
    "Manager Transport Engineering",
    "Manager Underwriting Engineering",
    "Manager Waste Engineering",
    "Manager Well Lifcycle Engineering",
    "Manager Werf Engineering",
    "Manager Wintel Engineering",
    "Manager, Change, Release and Release Engineering, EMEA Technology",
    "Manager, Client Engineering,",
    "Manager, Cloud Applications and Desktop Engineering",
    "Manager, Consulting Engineering - EMEIA",
    "Manager, Core Network Engineering",
    "Manager, Diagnostics Engineering",
    "Manager, End User Engineering",
    "Manager, Endpoint Security Engineering",
    "Manager, EUC Solutions Engineering - Nordics & Benelux",
    "Manager, FSP engineering",
    "Manager, Integrations and Solutions Engineering",
    "Manager, Quality Engineering",
    "Manager, Sales Engineering",
    "Manager, Sales Engineering - Strategic Alliances",
    "Manager, Sales Engineering & Training, Northern Europe",
    "Manager, Solution Engineering",
    "Manager, Solution Engineering - Financial Services & Manufacturing",
    "Manager, Solutions Engineering",
    "Manager, Support Engineering",
    "Manager, Technical Support Engineering",
    "Manager, Test Automation Engineering",
    "Manufacturing Engineering & Facility Manager",
    "Network Manager, Thematic Digital Competency Centre, Natural & Engineering Sciences (TDCC-NES)",
    "Operations Manager Technical and Engineering Support",
    "Operations Manager Engineering at Royal IHC Services",
    "Operations Manager Network Traffic Engineering",
    "PMO Manager en Manager Engineering",
    "Pool Manager Project Engineering",
    "Proces Engineering - Logistiek Manager",
    "Process Engineering & Continuous Improvement Manager",
    "Process Engineering Group Manager",
    "Program manager | Learning business partner Data Science & Engineering",
    "Program Manager Development & Engineering",
    "Program Manager Equipment Engineering",
    "Program Manager for the BSc and MSc Chemical Science & Engineering",
    "Program Manager International Student Admissions at Fontys School of Engineering",
    "Program Manager Regional Engineering",
    "Program Manager SmartReM BeNeLux | Engineering",
    "Program Manager Software Engineering",
    "Program manager Sourcing Engineering & Maintenance",
    "Program Manager, Central Planning & Engineering",
    "Programme Manager (Natural Sciences and Engineering portfolio)",
    "Programme Manager Software Engineering Curriculum",
    "Programme manager systems engineering & (digital) transition",
    "Project Engineering: Project Lead/Manager CMC",
    "Project Lead - Global Transfer Manager in MRI (RF engineering) domain",
    "Quality Engineering Associate Manager",
    "R&D Engineering, Sr Manager",
    "R&D Group Manager Validation Engineering",
    "R&D Manager Software Engineering and Platform components",
    "Recruitment Manager Engineering",
    "Recruitment Manager, Engineering",
    "Research Facility Manager Biomedical Engineering",
    "Riwal Global Business Process Engineering leader (BPM) - Group manager Accounting and Tax",
    "Sales Manager - EU Engineering Business Unit",
    "Sales Manager Engineering at ERIKS Flow Control",
    "Section Manager Eng. Techn. Inno - Simulations, modelling and Data driven engineering",
    "Section manager engineering (construction & Outfit)",
    "Section Manager Engineering (Systems)",
    "Section Manager Equipment & Process Engineering Lithography",
    "Section Manager Furnace ICN8 - Equipment & Process Engineering",
    "Senior Account Manager Quality Engineering & Testing",
    "Senior Consultant Systems Engineering / Quality Manager",
    "Senior Engineering and Technical Services Manager",
    "Senior Engineering Project/Proces Manager",
    "Senior Manager Data Centre Network Engineering",
    "Senior Manager Solution Engineering",
    "Senior Manager Solution Engineering (Presales), Analytics Cloud Benelux",
    "Senior manager Store Engineering",
    "Senior Manager Support Engineering",
    "Senior Manager Systems Engineering Europe",
    "Senior Manager, Core Network Engineering",
    "Senior Manager, Technical Support Engineering (High Touch Services)",
    "Solution Sales Manager Engineering & Geoscience Research Solutions",
    "Sr Manager Cloud Sales Engineering Benelux and Nordics",
    "Sr Manager Engineering & Maintenance",
    "Sr Manager IC-design Engineering",
    "Sr Manager Infrastructure Engineering",
    "Sr Manager Sustaining Engineering",
    "Sr Manager System Engineering DUV",
    "Sr Manager Test Engineering",
    "Sr Manager, Applications Engineering",
    "Sr Manager, Quality Assurance Engineering",
    "Sr. IT Operations Manager Subsurface Data Universe, Data Integration, GIS & Well Engineering",
    "Sr. Manager Solutions Engineering Cybersecurity, Continental Europe",
    "Sr. manager - Data Consulting & Engineering",
    "Sr. Manager - Integration Engineering",
    "Sr. Manager Development and Engineering",
    "Sr. Manager Engineering",
    "Sr. Manager Engineering & Projects EUROPE",
    "Sr. Manager Global Demo Engineering, Experience Orchestration",
    "Sr. Manager HFC Access Engineering",
    "Sr. Manager Media Architecture & Engineering",
    "Sr. Manager New Platforms / Technology International Engineering",
    "Sr. Manager Offshore Engineering (Geotech, Data Mapping, Survey)",
    "Sr. Manager QA Engineering, NowX Division (Process Mining)",
    "Sr. Manager R&D and Engineering (AMEX & EMEA)",
    "Sr. Manager Solution Engineering Benelux",
    "Sr. Manager Supplier Quality Engineering MP",
    "Sr. Manager System Engineering",
    "Sr. Manager, Principal Specialist Quality Engineering",
    "Sr. Manager, Quality Engineering",
    "Sr. Program Manager Design & Engineering",
    "Strategic Projects Manager West - Capital Engineering",
    "System Engineer - Technisch manager Engineering",
    "System Engineering and Satellite Manager",
    "Team Manager - Engineering Interim Division",
    "Team Manager Coastal Engineering and Graphic Design department",
    "Team Manager Concept engineering",
    "Team manager Data Science and Engineering",
    "Team Manager Engineering",
    "Team Manager Engineering and Consultant Water Technologie",
    "Team Manager Engineering Development",
    "Team Manager Engineering North",
    "Team Manager Process & Line Engineering",
    "Team Manager Project Engineering",
    "Team Manager Quality Engineering",
    "Team Manager Software Engineering",
    "Team Manager System Engineering",
    "Team manager Technology Development & Engineering",
    "Team manager, Project Engineering & Service Execution",
    "Teamlead Proces Engineering / Manager NDT department",
    "Technical Manager | Value Engineering",
    "Technical Manager Electrical & Control systems engineering",
    "Technical manager Engineering Shipbuilding",
    "Technical Manager Wifi Engineering",
    "Technical Manager, Operations Excellence, Senior Engineering Associate",
    "Technical Manager, Projects & Applications Engineering",
    "Technical Manager, Software Dev & Engineering",
    "Technical Program Manager - Platform Engineering",
    "Technical Program Manager, Engineering",
    "Technical Sales Manager Engines at Air France Industries KLM Engineering and Maintenance",
    "Technisch Manager / Lead Engineer - Engineering & Advies bij Stedin Hoogspanning",
    "Technisch Manager | Ontwerpleider | Infra & Waterveiligheid | Virtual Reality Engineering",
    "Unit Manager Engineering",
    "Wells something", "sales engineering",
    "Product Quality", "Product Specialist", "Head of Quality", "Head of Proposition",
    "Head of Securities", "underwriting", "Leiter Einkauf"]


    # Get the actual number of columns
    num_columns = sheet.max_column

    # Add the new column named "group"
    new_column_index = num_columns + 1
    sheet.cell(row=1, column=new_column_index).value = "group"

    # if you want the country in the person's label to be from the excel give it the index of that value
    # otherwise give it a string (dont forget the "", with the desired value)
    print("If you want the country in the person's label to be from the excel give it the index of that value.")
    print("Otherwise give it a desired string.")

    while True:
        country_row = input("the value is:")

        # Try to convert the input to an integer
        try:
            country_row = int(country_row)
            print("You entered an integer, well done! \n")
            break
        except ValueError:
            print("You entered a value by hand. \n")
            break
    # Iterate through each row in the worksheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        job_title = str(row[job_title_row].value).lower() if row[job_title_row].value else ""  # Assuming "Title" is in the xth column (index x-1)
        
        # Check if the country_row is a number or int
        if isinstance(country_row, str):
            country = country_row
        elif isinstance(country_row, int):
            country = str(row[country_row].value).upper()
        else:
            print("error with the type of country_row")
            break
        group_cell = sheet.cell(row=row[0].row, column=new_column_index)

        if any(keyword in job_title for keyword in to_check_ceo):
            group_cell.value = f"CEO, {country}"
        elif any(keyword in job_title for keyword in to_check_cfo):
            group_cell.value = f"CFO, {country}"
        elif any(keyword in job_title for keyword in to_check_product):
            group_cell.value = f"Product, {country}"
        elif any(keyword in job_title for keyword in to_check_cto):
            group_cell.value = f"CTO, {country}" 
        elif any(keyword in job_title for keyword in to_check_coo):
            group_cell.value = f"COO, {country}"  
        elif any(keyword in job_title for keyword in to_check_MoE):
            group_cell.value = f"MoE, {country}" 
        elif any(keyword in job_title for keyword in to_check_HoE):
            group_cell.value = f"HoE, {country}"
        elif any(keyword in job_title for keyword in to_check_VPoE):
            group_cell.value = f"VPoE, {country}"       
        elif any(keyword in job_title for keyword in to_check_cio):
            group_cell.value = f"CIO, {country}"    
        elif any(keyword in job_title for keyword in to_check_ciso):
            group_cell.value = f"CISO, {country}"
        elif any(keyword in job_title for keyword in to_check_cso):
            group_cell.value = f"CSO, {country}" 
        elif any(keyword in job_title for keyword in to_check_trash):  
            group_cell.value = f"trash, {country}"      
        else:
            group_cell.value = f"Rest, {country}"

    # Save the workbook with the new changes
    wb.save(results_file)

# Define the output file names
input_file = "2leads_with_less_trash.xlsx"
results_file = "3leads_sorted.xlsx"

# Call the filtering function
filter_groups(input_file, results_file)

print(f'File "{results_file}" modified successfully!')
